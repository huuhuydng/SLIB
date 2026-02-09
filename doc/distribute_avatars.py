#!/usr/bin/env python3
"""
Script phan phoi avatar AI cho 2700 user theo ma so sinh vien
Lay tu bo 7 anh mau va copy thanh ten theo ma so user
"""

import os
import shutil
from pathlib import Path
import openpyxl
import random

# Cau hinh
EXCEL_FILE = "slib_user_import_2700.xlsx"
SOURCE_AVATARS_DIR = Path("avatars")
OUTPUT_DIR = Path("user_avatars")

# Danh sach anh mau
MALE_AVATARS = [
    "student_avatar_male_01_1770392068785.png",
    "student_avatar_male_02_1770392084060.png",
    "student_avatar_male_03_1770392149564.png",
    "student_avatar_male_04_1770392165550.png",
]

FEMALE_AVATARS = [
    "student_avatar_female_01_1770392099496.png",
    "student_avatar_female_02_1770392115616.png",
    "student_avatar_female_03_1770392181132.png",
]

def get_gender_from_name(name):
    """
    Phan doan gioi tinh tu ten (don gian)
    Cac ten ket thuc bang Anh, Linh, Trang, Ngoc, Ha, Thao, Hoa, Nga, Thu, Huong
    thuong la nu
    """
    female_names = [
        "Anh", "Linh", "Trang", "Ngoc", "Ha", "Thao", "Hoa", "Nga", "Thu", "Huong",
        "Mai", "Lan", "Nhung", "Hang", "Phuong", "Vy", "My", "Hanh", "Quynh", "Chi",
        "Ngan", "Uyen", "Van", "Khanh", "Tram", "Ly", "Dao", "Duyen", "Tien", "Nhu",
        "Yen", "Thanh", "Cam", "Tuyet", "Diem", "Kim", "Trinh", "Giang", "Thuy", "Hong"
    ]
    
    parts = name.split()
    if parts:
        last_name_part = parts[-1]
        if last_name_part in female_names:
            return "F"
    return "M"

def distribute_avatars():
    """Phan phoi avatar cho tung user"""
    # Tao thu muc output
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    # Doc file Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    total = 0
    processed = 0
    
    # Dem tong so user
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[1]:
            total += 1
    
    print(f"Tim thay {total} user trong file Excel")
    print("Bat dau phan phoi avatar...")
    
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        name = row[0]
        student_id = row[1]
        
        if not student_id:
            continue
        
        # Xac dinh gioi tinh
        gender = get_gender_from_name(name) if name else "M"
        
        # Chon anh ngau nhien theo gioi tinh
        if gender == "F":
            source_avatar = random.choice(FEMALE_AVATARS)
        else:
            source_avatar = random.choice(MALE_AVATARS)
        
        # Copy anh
        source_path = SOURCE_AVATARS_DIR / source_avatar
        dest_path = OUTPUT_DIR / f"{student_id}.png"
        
        if source_path.exists():
            shutil.copy(source_path, dest_path)
            processed += 1
            
            if processed % 500 == 0:
                print(f"  Da xu ly {processed}/{total} user...")
        else:
            print(f"  Canh bao: Khong tim thay anh {source_path}")
    
    print(f"\nHoan thanh! Da tao {processed} avatar trong thu muc '{OUTPUT_DIR}'")

def create_zip_archive():
    """Tao file ZIP chua tat ca avatar va file Excel"""
    import zipfile
    
    zip_name = "slib_user_import_with_avatars.zip"
    
    print(f"\nDang tao file ZIP: {zip_name}...")
    
    with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Them file Excel
        zipf.write(EXCEL_FILE, EXCEL_FILE)
        
        # Them tat ca avatar
        for avatar_file in OUTPUT_DIR.glob("*.png"):
            zipf.write(avatar_file, f"avatars/{avatar_file.name}")
    
    zip_size_mb = os.path.getsize(zip_name) / (1024 * 1024)
    print(f"Da tao file ZIP: {zip_name} ({zip_size_mb:.1f} MB)")

def main():
    # Kiem tra file Excel
    if not os.path.exists(EXCEL_FILE):
        print(f"Loi: Khong tim thay file {EXCEL_FILE}")
        print("Hay chay generate_user_data.py truoc!")
        return
    
    # Kiem tra thu muc avatar mau
    if not SOURCE_AVATARS_DIR.exists():
        print(f"Loi: Khong tim thay thu muc {SOURCE_AVATARS_DIR}")
        return
    
    # Kiem tra so luong anh mau
    existing_avatars = list(SOURCE_AVATARS_DIR.glob("*.png"))
    print(f"Tim thay {len(existing_avatars)} anh avatar mau")
    
    if len(existing_avatars) == 0:
        print("Loi: Khong co anh avatar mau!")
        return
    
    # Phan phoi avatar
    distribute_avatars()
    
    # Tao file ZIP
    create_zip_archive()
    
    print("\n=== HUONG DAN ===")
    print("1. File Excel: slib_user_import_2700.xlsx")
    print("2. Thu muc avatar: user_avatars/")
    print("3. File ZIP (de import): slib_user_import_with_avatars.zip")
    print("\nBan co the import file ZIP nay vao he thong SLIB!")

if __name__ == "__main__":
    main()
