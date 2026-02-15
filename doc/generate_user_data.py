#!/usr/bin/env python3
"""
Script tao du lieu user ngau nhien cho SLIB
Tao 2700 user voi thong tin ngau nhien va anh avatar AI
"""

import random
import string
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl import Workbook

# Cau hinh
TOTAL_USERS = 2700
OUTPUT_EXCEL = "slib_user_import_2700.xlsx"
AVATARS_DIR = Path("avatars")

# Du lieu mau cho ten Viet Nam
HO = [
    "Nguyen", "Tran", "Le", "Pham", "Hoang", "Huynh", "Vo", "Dang", "Bui", "Do",
    "Ho", "Ngo", "Duong", "Ly", "Phan", "Vu", "Dinh", "Doan", "Trinh", "Luong",
    "Mai", "Truong", "Ta", "Cao", "Ha", "Chau", "La", "Quach", "Lam", "Nghiem"
]

TEN_DEM = [
    "Van", "Thi", "Hoang", "Minh", "Duc", "Thanh", "Quang", "Ngoc", "Hong", "Anh",
    "Quoc", "Xuan", "Nhat", "Gia", "Kim", "Phuong", "Thuy", "Bao", "Dinh", "Khanh",
    "Trung", "Hai", "Duy", "Tuan", "Huu", "Cong", "Vinh", "Tan", "Phuc", "Thien"
]

TEN_NAM = [
    "Hieu", "Anh", "Hoang", "Minh", "Dung", "Tuan", "Hung", "Cuong", "Phong", "Dat",
    "Huy", "Nam", "Khanh", "Long", "Duc", "Tai", "Quan", "Binh", "Hao", "Thinh",
    "Phuc", "Tri", "Loc", "Sang", "Trung", "Vinh", "Kien", "Trong", "Khang", "Doanh",
    "Tien", "Quang", "Bao", "Nghia", "Tam", "Son", "An", "Danh", "Lam", "Hung"
]

TEN_NU = [
    "Anh", "Linh", "Trang", "Ngoc", "Ha", "Thao", "Hoa", "Nga", "Thu", "Huong",
    "Mai", "Lan", "Nhung", "Hang", "Phuong", "Vy", "My", "Hanh", "Quynh", "Chi",
    "Ngan", "Uyen", "Van", "Khanh", "Tram", "Ly", "Dao", "Duyen", "Tien", "Nhu",
    "Yen", "Thanh", "Cam", "Tuyet", "Diem", "Kim", "Trinh", "Giang", "Thuy", "Hong"
]

def generate_name():
    """Tao ten ngau nhien"""
    ho = random.choice(HO)
    ten_dem = random.choice(TEN_DEM)
    # 50% nam, 50% nu
    if random.random() < 0.5:
        ten = random.choice(TEN_NAM)
        gender = "M"
    else:
        ten = random.choice(TEN_NU)
        gender = "F"
    return f"{ho} {ten_dem} {ten}", gender

def generate_student_id(index):
    """Tao ma so sinh vien SLxxxxxx"""
    return f"SL{str(index).zfill(6)}"

def generate_dob():
    """Tao ngay sinh ngau nhien (18-25 tuoi)"""
    today = datetime.now()
    min_age = 18
    max_age = 25
    days_ago = random.randint(min_age * 365, max_age * 365)
    dob = today - timedelta(days=days_ago)
    return dob.strftime("%d/%m/%Y")

def generate_email(name, student_id):
    """Tao email tu ten va ma so"""
    # Chuyen ten thanh email format
    parts = name.split()
    if len(parts) >= 2:
        # Lay ten + ten dem viet tat
        email_name = parts[-1] + "".join([p[0] for p in parts[:-1]])
    else:
        email_name = parts[0]
    email_name = email_name.lower()
    return f"{email_name}{student_id[-4:]}@slib.edu.vn"

def generate_phone():
    """Tao so dien thoai Viet Nam ngau nhien"""
    prefixes = ["032", "033", "034", "035", "036", "037", "038", "039",  # Viettel
                "070", "076", "077", "078", "079",  # Mobifone
                "081", "082", "083", "084", "085",  # Vinaphone
                "056", "058", "059"]  # Vietnamobile
    prefix = random.choice(prefixes)
    suffix = "".join([str(random.randint(0, 9)) for _ in range(7)])
    return f"{prefix}{suffix}"

def create_excel_file(users):
    """Tao file Excel voi du lieu user"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Header
    headers = ["Ho va ten", "Ma so", "Vai tro", "Ngay sinh", "Email", "So dien thoai"]
    ws.append(headers)
    
    # Du lieu
    for user in users:
        ws.append([
            user["name"],
            user["student_id"],
            user["role"],
            user["dob"],
            user["email"],
            user["phone"]
        ])
    
    # Dinh dang cot
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    
    wb.save(OUTPUT_EXCEL)
    print(f"Da tao file Excel: {OUTPUT_EXCEL}")

def create_avatar_placeholder_script():
    """Tao script de sinh anh avatar bang AI"""
    script_content = '''#!/usr/bin/env python3
"""
Script tao anh avatar AI cho user
Su dung API de sinh anh mat sinh vien
"""

import os
import requests
from pathlib import Path

AVATARS_DIR = Path("avatars")

def create_placeholder_avatars(student_ids):
    """Tao anh placeholder cho user"""
    AVATARS_DIR.mkdir(exist_ok=True)
    
    for student_id in student_ids:
        # Su dung placeholder service
        # Ban co the thay the bang API sinh anh AI thuc te
        avatar_path = AVATARS_DIR / f"{student_id}.png"
        
        if not avatar_path.exists():
            # Tao placeholder image
            print(f"Tao avatar cho {student_id}...")
            # Placeholder: su dung UI Avatars hoac DiceBear
            url = f"https://api.dicebear.com/7.x/avataaars/png?seed={student_id}&size=200"
            try:
                response = requests.get(url, timeout=10)
                if response.status_code == 200:
                    with open(avatar_path, 'wb') as f:
                        f.write(response.content)
                    print(f"  -> Da luu {avatar_path}")
            except Exception as e:
                print(f"  -> Loi: {e}")

if __name__ == "__main__":
    # Doc danh sach student_id tu file Excel
    import openpyxl
    wb = openpyxl.load_workbook("slib_user_import_2700.xlsx")
    ws = wb.active
    
    student_ids = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[1]:
            student_ids.append(row[1])
    
    print(f"Tim thay {len(student_ids)} user")
    create_placeholder_avatars(student_ids)
    print("Hoan thanh!")
'''
    
    with open("generate_avatars.py", "w") as f:
        f.write(script_content)
    print("Da tao script generate_avatars.py de sinh anh avatar")

def main():
    print(f"Bat dau tao {TOTAL_USERS} user...")
    
    users = []
    used_student_ids = set()
    
    for i in range(1, TOTAL_USERS + 1):
        name, gender = generate_name()
        student_id = generate_student_id(i)
        
        user = {
            "name": name,
            "student_id": student_id,
            "role": "Sinh vien",
            "dob": generate_dob(),
            "email": generate_email(name, student_id),
            "phone": generate_phone(),
            "gender": gender
        }
        users.append(user)
        
        if i % 500 == 0:
            print(f"  Da tao {i}/{TOTAL_USERS} user...")
    
    # Tao file Excel
    create_excel_file(users)
    
    # Tao script sinh avatar
    create_avatar_placeholder_script()
    
    print(f"\nHoan thanh! Da tao {TOTAL_USERS} user")
    print(f"File Excel: {OUTPUT_EXCEL}")
    print("Chay 'python generate_avatars.py' de sinh anh avatar")

if __name__ == "__main__":
    main()
